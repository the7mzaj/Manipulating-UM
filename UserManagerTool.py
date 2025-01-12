import os
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

class UserManagerReport:

    def __init__(self, root):
        '''
        Initialize an object with browse button that functions to select the relevant file, and generate function for the generation prupose.
        '''
        self.root = root
        self.root.title("User Manager Report Generator")
        self.root.geometry('500x500')
        self.root.resizable(width=False, height=False)
        self.file_path = None

        #Interface

        self.label = tk.Label(root, text="Select an Excel File")
        self.label.pack(pady=10)

        self.select_button = tk.Button(root, text="Browse", command=self.browse_file)
        self.select_button.pack(pady=10)

        self.generate_file_button = tk.Button(root, text="Generate file", state="disabled", command=self.generate_file)
        self.generate_file_button.pack(pady=10)

    def browse_file(self):
        '''
        Selecting the xlsx file to generate
        '''

        try:
            self.file_path = filedialog.askopenfilename(filetypes=[("All Files", "*.*")])
            self.label.config(text=f"File Selected: {self.file_path}")
            self.enable_buttons()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while selecting the file: {e}")

    def enable_buttons(self):
        '''
        Enable the button whenever a relevant xlsx file was selected
        '''

        self.generate_file_button.config(state="normal")

    def generate_file(self):
        '''
        Manipulate the xlsx file to the desired report
        '''

        work_book = load_workbook(self.file_path)
        work_sheet = work_book['Sheet1']

        for merged_cell in list(work_sheet.merged_cells):
            work_sheet.unmerge_cells(str(merged_cell))

        work_book.save(self.file_path)
        work_book.close()

        #The following will get rid of the irrelevant data    

        df = pd.read_excel(self.file_path)
        df = df.iloc[5::2].reset_index(drop=True)
        df = df.drop(df.columns[5], axis=1)
        df.columns = ['Username', 'Name', 'Email', 'Brands', 'Roles']
        df.to_excel(self.file_path, index=False)

        work_book = load_workbook(self.file_path)

        work_sheet = work_book['Sheet1']

        work_sheet.column_dimensions['A'].width = 40
        work_sheet.column_dimensions['B'].width = 40
        work_sheet.column_dimensions['C'].width = 40
        work_sheet.column_dimensions['D'].width = 10
        work_sheet.column_dimensions['E'].width = 110

        work_book.save(self.file_path)

        messagebox.showinfo("Action", "The file has been generated succesfully")

if __name__ == "__main__":
    # Initialize the Tkinter window (without withdrawing or deiconifying)
    root = tk.Tk()

    # Start the application
    app = UserManagerReport(root)
    root.mainloop()
